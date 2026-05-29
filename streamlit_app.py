import csv
import io
import re

import openpyxl
import streamlit as st

# ── Logo (embedded) ──────────────────────────────────────────────────────────
LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAMAAADDpiTIAAADAFBMVEVHcEwAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADR7hC8AAAA/3RSTlMA/gH9AvwEA/v6CwXWTeE9lYLv"
    "+aQWFwYYGvhvh2xtBzYZhgr3blEJ0fXuCPP2Dd8MDvAhhckTIiXt7Mul4B0V6jUS1/HYMhEv9PLdwuJ6"
    "mB7odopxa2pMfcwciKnms0QU5Sah5+Trc9vSWRvNxtq9HylDw9SEmzMkrDdOSWi6dA8ogJQ4LS6xWprO"
    "ucqwvuM6UnLeQiqMZ5zpI09Asnk83Lg/eEuRpjGNyJ9pEL/Pq5aPvLVcRsFdMKrQJ2OJqH9BNGFQO7Y5"
    "XlNw2V+3e4HVV8RFVZ6uk5Igl5BWZqKvKyytx0iLnXenjrQ+fmJUu3xkR6OgW2WD05nAxVhgdUqxoMEy"
    "AAAgAElEQVR42uxdCXBV1Rk+6z338sBSnvAk1MaZRMOSBcKWBhsgEMIWFqkSIMNOEtbKKhh2l4osAy1a"
    "BBELbYGGVVsoUKCKYB0qlVFRa4vF0nYcpGDtdBx0enrOufclLwuQSoDk3v+bCS+85L28e/7//P/3b+ci"
    "BAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AEDdAvEAKxE8yTPGhGAuPA3QTwoP5llQDf9JPSpeUvF5YWkIRmJsAilXDYDPkJr1zImhr999adu6d97d"
    "OHLnri4uVmyfunzlxoJ1y86/8OOLM17JaNogIowigB74YPszxNpmPP72op/NmzlxTF8sJaXyqsDqZzQu"
    "vV/H4+NGPDi09/AEgkgF5gALWo8Ej0ho/eZXD6/smBbnVBA6xphSyq8CSjH2fjG5x6rP3zzWa7LF9Fuy"
    "ih4EUJdhTX70070Li7/DMae40i73UL0JiAHVtkL9Wnhwl4ILk55ItZRKaR4Bq1sH97v3ENLS6bSpdN2K"
    "QiU8zrHUcsQVIGsKVwfMy5W9SBm8cM8vt6Yqn8BACeoez49u/PYfPjtzsN70WNtyWiZvI0Ts7nBt/W3b"
    "vroLcF8Y+0puO1yxA5mYt/A/Q7sSN5RE4BDqCFzfTD76at2uPpKW8TlcdTNf1dxXthBV/xs1BdhucnpE"
    "o6YqQmBACusM4SPNJ+2ZFbvjy0C549i8TPI0rsmEhl1eOjfui22nTs0uKr284KHS0tItS7+59IM/r9u4"
    "anvO+GSniuLEagW3sXpP2nPmiM8Wa3cASnB7uV5ERXlJJ47uTK7OtVNP8FimFN7x/SPbHtg/9J72ufHX"
    "cOBEJHU9tPq/v9myd972tCYxXqACqE2lsi/dzt37TITEOCDALYdAzRrN+ZEnpGponB0evGvlkAW/6tWV"
    "xKI8M+hCxGSCWTTcjzTvVfLDxi/P6plYnaPQyqW4wbAOqxuAGG4bsvYfmGCE7/K2itLPPHjg0pp7OimK"
    "EGWKrLLBrlIU8moF0ZqAsgiINVv91uFPpiVri4JjrEHUM+CcxvOzCQJXcMtZ/+TLI9M9S1851k+f8k7p"
    "5ofjy0Ravq2/5t8TDdo3GnCgR7qROHd4TITAbe6kDRyVYCHIE91CDdjx1u5u0mzIymY/b+OAGQ8LN3Vn"
    "hSxRi0JJyFgze99gQyowL482tC2wixv/1oJs8a3CjIF51aZuiu8vfTJbmMCw3MHXnkjMO3Vdu+Xs+DCu"
    "FE9qfWi4dLowOgCW4KZiff6sRFlx7+tvx6+8t1eqiQkty4o68trdjzoRrN/Q2rE5f19rwwncT+GlGXB4"
    "4clcxUyhjHjTLL9IOnauVVW6H+7yrVGpbjE/Wve/6SLI+t2eHraUdtihRgHdVCPtduYRC+pGNwud83tU"
    "sLw6pyubvNxirGH6t9D0ElMfjHTP79jTDQxiso9458lUREIRAWaglvHa/emVtz7vtvup4Uh39NS2v68p"
    "lhT1T3FVwJW+epR8TIdDSIRCAjSg9sjX4kkfV2b8uPW4i52E55pvW5snYYOK9jk6GOAuJzCfsu+814yh"
    "ABWoHdw1IM2u0LdDZZ/lJ5tFtPR1T9/t+2RM/XGBpi8a5ig+YHsJCcqVcZr5dpJWARDejW4wgpp36FY5"
    "xdvyjfbMZfzi9tMt9SkQWb3tRSV1u7zwhHmPy0nEkBMQ441g+geZ2ImW47Tw+eCB31DunoV0Q2/dWFxm"
    "KSeUet+5OFOBKk8UTivKVTYC6OCNiH9ZCrW59MJt3bPZ/9XckLG9QtSdWMvEnQIduvCi9gQ2d/kK5/K7"
    "ixZDvfAGOPbnfXBMQye2M5c9Uke9qsdBG2z4p6NdgWYDigpwTosXtYNZk6/p+5eFbSeG/MmcRboTR9TV"
    "JIuWMmPo140zy/KT+iGc9r1sBmbg/wz7EBq+rZXk5WV+bO/7MN7k2OpuolXXAVRUwLr+vZ+k1OFuiVr9"
    "02N/vAC51nwdBRGpRydQj/UZ+bcq2MRiLG0dtlxM60C7/VO4jDYpUBvzlj81RAGEW7NVDJ1M417zhRF/"
    "n8f+wnSVpb64L0VQSclpGVUBHFZX81L3CLiBGmLz1GgnnrYBNHlbRoggy6o/BRbTLEpmHNeBgGcFFCH8"
    "xSAUDxMF13ejc4/onsvo9uet3xsk1J4i9e9CCBq1XPcrmdyQbhnoOaAt9IxcDw0uJKpQOlpml+nns0zx"
    "TbD6p8paa59fSKXbskZ1vXjWMQatY9dIp1io0USMy0O/VnMy4pnp8KinsSyLF0OvSG9AUXcS4x8s0U+D"
    "rKvH3H/o7qqyBt93t3oNPvU4ndEGiS+fNklBr2EgeUAEMagUV4u706U7n22c5pXnFZeq9+ZSD6wmPJin"
    "tr/LaSmVu9bqbBZYgUoLhbbO1I11ZqZTrVLxxZA/pm5MT8Do88nSkEGlAcrF7emkIkWQeQUknErxZncV"
    "b5aZl1KV8ffFJlE2jESYGHtOqbYht1rD7yjRY00AL15iIvLKFJcouRpQsB6JiE8WyHVihKGhByk2GoCp"
    "Lem4bGSBCrgKoDZDBxsr728666QzcQMxR7z4xbd5ioCSHsiU2Gsh5rLffGRZQAR0TCS2fiyp21On0Lqo"
    "DfLTbAUpzwugjJ97h1SZCeM5CUgEPiegHH2oRR9p6z1hMDJDLYqfNkbM6KmS9pp+0hsrojTccBOxICC8"
    "a6RyjaaVjnL53FO6w863a0JYpO0fE3nZVKGdL3SOM9A6UJInOY0m/v/VlDB/J8vVpT1+0JZukZDa9qr3"
    "mQisAihJhw7b0jETFWr7521wz/vxdbVEXVzSkBTJPSZgj5mEgusFSOedUtrc9Yj4zlwUiFO81fVtbqkV"
    "3nUD+EwCCWowsKGQeo30jlP4JULBUAAd4rbZ6+UEKMd0ZmcUvPkRXRE7ZaojWOdInU8GxZzh4/uLFxb6"
    "LE3GuakvSgtLgtUrRIybH32ce+P1XPYssggLzm0c9Dn1KPtOnfpyO5/sNwPVJsIsFGKP5kTP34vDB7uj"
    "4Gx/Fzr4+1Nr6pjMII3Du5sGyg1Y1uuJ0p2gwtz5WyoKERSw6Qmd7ui9i4fd5hcHt1wiQoHhgoQNkQ73"
    "pv1STpoZr8BNz2g/EH+Ghh1v4DmzJCCXrcS8eB62vZS4bPlkUPvl9RAJWZDsHi+hogHnDRGApSBCoPdX"
    "RLs+HVqwQ5v/oIJZoScaemeLUJyyR7mFAMTAmybQOG/eh7dAwe6MYRbL3u2dbK5I8ch2BPm7Pqhozn19"
    "semSVdfcba0lhAj0OSq6NXCpO/iui2FT5lrI15UQi3y7/AYew8bqZiAR7KYIIgha08TrF4yTad1RGz9f"
    "LZpN3fBPqfyRXIIQDEnocdLe07zqkI0L/+pjGkDQF7LsLk6z40nZSd5B1wDS7CfSdofgZdy//bsokbPS"
    "3MdJa8BXwoIhuSgEC43j2NylTH095Es1J4jknpVh7zTFzLUiBOIvp4JCkKPU6xGQMt+fJ4t1uiLDbvVL"
    "NvwIwQF6FfaHFRKfOtgUh5QGzPYhNwoNmqrlb47Vfro5WP8qiEfznzMnzGFuy8d8pgGK1gzKkXH6Rgs4"
    "kS9fLGBIvhouKE6kyURquqPoQF9pABEo6w80zr29izwQL+DwzGqJABvbULcImBsU+koD4lH7HH2+r3YA"
    "4fcYIxD9Va8BqPkuaaZjpdYAH1nJzsr+U/e+W8/CHVWuHikRlLWKmnyQ0QCfXJaSf7Hhtno0+gWQ/zVV"
    "QKQu1GulRyXoQOITBWhaLN3mN5B/zTSAmljZHxqgBD76tIy2/oP8r79eiBVoDdApU19EgyS7vx2dgQD5"
    "18hjWloDpI0px6fqd8JM09g2v/fOycP/Y+/ag6sqzvg+zp5zLpcg5FYyhCgjEWJMKARQSIsIqBhbZIoK"
    "IhY6IqNlUMCKDxQrPjBiY7ANCoIi9YFV66hVSx2Ftgr2ZaX4GB10sJaRilLbWqfttON83d1zo5ibTe45"
    "97zu9PvxD0wOJ3v2++332P32+/h+lH9xQbNkgLouIV1msamsGaDCvY9pZ5W86Xj2W2Q4mLV3aj9Azpu4"
    "psx9msdlSMNVYSS+CeVf/LytbM/7AUCHk7JNmZEWYLbXP0WyYDoWy/bjCPRp9/wAgBFry1iXPeKqbDfO"
    "M/RMrJDqZ+Isq+kcVVxSmk8+aRUr18SZV3P5ql/wBMrfJwHYxGXgqvwp7p4ysBxLCMgR33YyqIZP0gh8"
    "k2AA4NN6Mmt8B3X19fFce11Znp6N6chARrVWpwccG+XvN4CyrB3HyvmT64fDhDJMEXJW/kkXf5Hybxvl"
    "YA/VIAz4e0bVT1AdsmvKzYZJeW/XHRQlfTcPYlgLLRAD2FZ1fsKFK+jz5XU47FSRAaDvf1J6wR9ZFsUZ"
    "hAFO1rmQqv0AGQzm7rLKqd2MZX3SeeFtxNOY/hVcCWTf0DqAu7BmXPm4/4xlp37NS/8F/jarQ0kGV6Vk"
    "ud4LkAx4qEwKC+tCDyfuypdAg7EN2D27JDNw6kJ1jk5dl64uC09a63unlbqq/BcV2/NbAojA7vSqGaq6"
    "PBXVMKwsfGk5xiE/p/kbYOc0qdNMJEApC4p8q15oR1DUf0BS3z1N908+W1DvDviMSxjDI6CSF9SRRwvl"
    "CrpwykDHSnk9DdXa+807O8veXG0RDAFKJwBZrx1BIWCvOiVIuwao+sgLXan4tIpg1/QwjEDdbq0ChCum"
    "q35j6XYAydB87yfYQtADDIkBGxZ4VXVFZi1LsRFQ9X+zG3WLVKBw0Y/x/k9Y80pWeO20gC8+PMVLSnl8"
    "39is3X8KlVdga8TQYsFs9q/expo6WbfSS1T551a9BUwFPMfwCCg0AjjEafVu1kBuXWpP1pV3MksfYALA"
    "amKh/xemEVjaoWsrADQ/mFqaMtLSqB1ACl+dglIL2RG8X0UCXLgwrSqVC0uFJ7W7VfoKFfy0F1FmIasA"
    "h9UILxTkA9Ipf6kArlfNv1UF2GG4/xM+Bex27hkBcVU6DQBrqdflToG/i+KPwgjcfqwOBTJw3CiWRgXA"
    "poE+BMg0L7UwBzCCNcbe5jmqt4TPtdMWDKrS9y/rg2tJ0f9aeAkgikVGso95jiAc/WLKiqyrCGBqvUpe"
    "kh7AwwzbokeEiXP1CuNwcVXaIgDGluj9H4DNR6GgIsNWXUSMu3x/uvwTiZlc6AZobgWKKTo/0PrYy7QQ"
    "o1elqee0HMnEBboOEKfv4Q5wdJEgIXPWcF1uofolO007rcz6EuSU/Yc1g7RDgIiKAluFrrbOxdYUEYCR"
    "q3UdKCrcp4iNKQBRhoKk1fV87Y7DUxNrM6fPMaqWAc/AcrUDgNtAUfoBz04CoY/cH0jNPFtkdjUVQhKg"
    "+Vlmo/yjJUD2WrXYVJbwjtRM9PhJqpaBJOZ04mAh4KgjAXIpVXW3KXyUkgpiVnZLvhFkW1ZlBaGUovUD"
    "nVdc79ol3JwOd8vaAboUpMhUqBAQDUDEDLCt9brylqALp6RitVlLeE73u9quLjMiAaKGY8+Zq5vOU5id"
    "/GxLm7RVX2B3Ye51Uv7oAcagAshv9JRzuHNc8tPtrDxJlQGRdLye2ZgHEk/YPWSZVgECbkhcBTCyX6g0"
    "8AyM/DLmgcZmdc8Gb+fttJsS3na12FFz9T2wEfA8yiVGvE5dHQpuySbrB9qkxtuUgEuxEWCcmJwDHQry"
    "w0iC2RdS+yw91ksEz+xDocSIKtKfei2YW+sSLMHFmHWmzgPjMA+FEu/Su3009bqwr0jQD2TszUqVoMJh"
    "xFW4AxgzzvWKMMCjVUny8H3Q3Y3gMYJJADFjznn5RtxJpmBtaAa9Ld14BW4Axq18yX2gU3Dg4uT6sLM9"
    "oDMB+XpUAPEz4MSRavK5y19IbBDzm0GouyCNDyIBErC/Z6l2PCIDLyU29z8C7y745QRvAsVPAOvUE9Ts"
    "UxArkhrDP743b+fOA60HLiEYAiRBgYPtt999997XlzyS1Ajy2T8MkwAS8gPVwmMJNmPRZSCZvhiC8kjI"
    "EUh09lkeqAASJECSJ7DsM6AokiRA8gxASSTtDCAQCAQCgUAgEAgEAoFAIBAIBAKBQCAQZQI8ZETxIwGQ"
    "BEiABNdeggJgqAEQCAQiEdVrfQFhK+NeTUznEByFLkNAg4BAoJOJiBKfad/wVbDW7Y5j23bWg/yb7f2O"
    "Q97tPaSesr2ff/5E4RC++L4uyL896LB7uuvIQlsOh872obCCDFy+zvnCbHTO7+fzV+J4/5+WgpP0RISr"
    "Iot+UcWwmR6OKEBTCReA6/YN3/irwY8UvHPWNWf9oMujc2bl8fux28ZeNvZzvH9l4VfUnbFt6NihZgwM"
    "PIP2h+r/13SDpwPchLaGNI1f+9bBI4vBzL+9sGNRk2Mx4tiWv1+1scaAwcVWkmN3gOtyBVqAX5dAwF9A"
    "96AufazLo7cZHgX3hcKSiCv7gehmqJ/hBitoDb3saNM4hhZfl48pbU5G3Tj4/OOaBfgAPf3Rx9fd3iAV"
    "kePn5u9q0/tOKrqG4z2CCldBFGBCCRpgD3czghfCzVW/2nUE3AC4nxRMxsrFNJMRZlTfRQJeXG2YkZPI"
    "dIPZfro0sKaND3e4WqaieMiJUYy/5YZ7a4kfP6a/a8CuYofMsovza6eQRQtKaFh8S7dvVAXPL+jaZfIw"
    "E4vF2ZbTlYN9+qkiZT2gzQrYxK6hUZOu68qUeNkHAaY+cQIH1UtRS7UnXdUF8mH5Pyj0vWyM6hRWJAX6"
    "0y7S6/x336KHzN4AEwHc4UEJwF7sXvxKeM/YxRIAKgo1QG0/r0iZSZNyOCNg8Zq6StNL/1kkARj5YF5O"
    "ThsHr6Uy+ES+R2zl5depvsxFmZ3+pncVTwBprU0EgH8HVgA1hjdKAf2OFEsAt6KwH2Gfngkgcd4lwUxA"
    "baVhzDCgSAL87En1gXLxcwgMRQF68nvPsuIczzAIMOrrRgJ0NAQNqS4yTSZULoqWAALeDeYG9imFAMxi"
    "TXs86ysolARdhzn3aQNxrHgI4Kw2EgD2BSpIyMg73PRG3kqKJgCtIP4JQAX/vh0zAdQg943URp+XKP48"
    "Bzi96JfE6p0CYRCArQCjuB4IUguM2WSTUQHArOIJAEEIAFz0nRLECJREAOdMap7FIBQQtPE+0ntIGwYB"
    "yPH9XNPQO5oCbanU7QbTVDTP90WAgs/onQDy53fESgDpsw9cAi6ECioy/Jnje1UBoRCAnA9GhX1vgKl0"
    "tAUwDOy7JHoCCDE5wLCDawCn5dtUSO+Dh8wB2jamt+8IhQDsD8LoBGwL0JzCIvvBKKN10RNATt00y/+m"
    "enACHDZJCZ8ChE0A+G1LL7ta4RBgygKjG7irNohLvduoAEaMj4MAkFlnWfEQQIpn7Wkqug1d/ro/0JrJ"
    "Pa/AcAjA/mzcCqA3BrCmV9YbCTCNxUIA3jwmHgJIN23yVyA8568ACZw+v8dYIBwfgNwMxu3g6wMogCPM"
    "EzKsm7mMgAA5WB+TCXDGzAUaGQGUJA83GwEWEgHYuBmUd88A3uY/EHSeNMYA8I4TiwagQF9hMRCAkdr/"
    "AI2SAFz8pMF4PKgIQMMgQHaviQBUTPXtTm2oNxrEh7ozaH4JUNwx679GxUAAi7wGOQ4RKgAu4A3VntdI"
    "AB6GCSAHgZuMwCbfEdXzYCTAhyQuAlTDdGZHrwGuES6P0gCo7mz0L5Zt1gAiFAJMbDQS4GJ/ZUnls63G"
    "FcFbWFwEoG7jTdFrgPGNAno5ni5dB/C5pu0AKywCMOeHJgIIGbj5I8DSSmEiwMKG2AggVeMB29fIfRNA"
    "OkfLwaVRGoC8QzPBkOIXGgEIm2mIA6hwr/WTqSgffQqM52FjnfgIIMfwia/Ghr4J4JC3clx+a8TyV59y"
    "ddQEIKvqDQSgtN0ungAqE3yCeUncQ+IjgJy2kxZFSgAyZJmON6ImgCRZWzZqDUB+avIBxNHz/TgBbNAM"
    "45QcU8ViJID0a/tH6gNYF6oToMjlr3Mcnut26YSoAci1nHZ/nClgsC8CDDeuCfcJFi8BOHwnOgIwcupm"
    "dfxDe8rsKDoFpKdHVYbIwiHdmbMwCdDiGvLYBOz040ux1/5H3bkHR1WdAfw877nLhgYTzJIgDUiSCoFQ"
    "guFNiUHAEhAmgMozigR5SkEUBqgVkVqGh8BALVYtIFpGg7TIUIqvIuNQqVFbJZ2ptFKn1seobZkOaF+n"
    "525CJMn9zt5dzrlLzh/8weTu3nvub7/39x1wT3AmChUATPlv84ktABg5EdEX/+CAxSHxSkCIgAY4KP4M"
    "Re0CgMqkf0iTykETAhMgxK01kADAc/uGC4BH72uWJIB6ELE/gCeiIExcER7/VqCW6EKV727HsgRAdYAE"
    "UP/1aWAAYuhmWCiuJyRsAGRpP2IHACFWekEaTSqXjnmwquhYboC1cvVrp+YA8YQLALhHLNsAbGWe9I9p"
    "cfxYYHcqip6GAWiPwgYAc8+FtgTAPr39N+7GHCUQA3/7rc+86es+N9X9X2UZAJK/0FNavgD0zA3caNZu"
    "LrgtfSpCB8ATar9mVmwA8Y3zcBUYdml9/2jr3njdQrFb/AhoAmCeTwzFqApAV2N/ACjlhwLrgKUuKABu"
    "J+EDoLRwRoVjxQjsisEbwfT6+TESJUkBoNaWGRoVgLfYBuCBiPQFQGHxy6D9iqwDLBe3ojTYANzF0wSz"
    "AUCdxgHsmSmSffnxNWsYhwDIlpMsG4Ei9ibgjHCcVRIwFBBbBgIwrCIdAKjb6d21mJkHIPYxhu/jZkZS"
    "W8tLW2LVJAHkc62jgUYBcEQ9998A5aUeD1gWsh20i+lT0XSoAG8tjgUSYMkB0LccAiCP1rHki5Ibfxyn"
    "gfw+5XJEd7sAMO9F+Pqi6svfCdYdR6aBb8E9gdIiAbz1AokaB2BUD6gUQ5YVp3KwZ3xzBNoA1oXkrbML"
    "AGElu/wBULsyY3QQCUCi80AJMCcXkTQBgGf3M68C3qIgAO8SJ0UA1GV7QADwXrsAIIesh8rtKF8V6CPO"
    "gHaRNxYkXQBQeU8QEZAcAI8Df4vlyGKW0tG+jWaAf1MVVeq50DIAglW7UFc/7hBIq70Bv4S/wzOQrKsA"
    "LlcHsGGSA2AtoC6pnIZSA6DxB7LNN4VAOcf7LAPARP9xEsK6z+gAj5RfBodkH0kjABQP6WsagAOgBviH"
    "SPF094YNOhaJJ4daj9aRZy0DQAR7GnLiqKwO8EjvwRrAK89KHwAR+bcYMwvABgAAnN2PpAhAww4Vz434"
    "DJahPCI3WwYAOawIBuCHiZ9JVMF2WBdE7AOA4cxM7x0s0ei85AAogyRA5WiU8ri/+IV/vPPZa1qvowvv"
    "/MI2AIJcWQOlcj0vNNFDxe4C3+S3l9gHINulMBh4ppNIMScHwEYIgJPFKLSRuIYBICz6koQAwNWJzEDy"
    "XgEog/cj2wBgOXOxpjhH0vmJQhnJATAF7Mjo3oYBQPNB00bendAP6Axdmicn6Q6FMiQBxm93dRHhyhxh"
    "EoD3oZ2q6d+GASC5w8Aih6MJn+ogdCl1z+iuMwGAuu2h7KymQo/TPwiTKmA3BAAeSNosAISIl0ATi+5J"
    "8FgreoG7fxcKBYBHZsBdOi6OfGJSAgwFATjUlgFAr0DeDZb3J7CinoEF8BPakWfGACCn4TEtnNNF3Yk5"
    "APaBALzTZlWA1+vyrVIwFr0xwWPthgFYLkKRAMxZ7Ooqr2/KNwdAFxCA8uFtFwBG0Mf+QlRtMP++Np46"
    "YDDUESqfdcJRAQ4apZ3VlbWDGQPgOAiArCKsrQKg7ruLf2eftzf3a4fwvhIBI/F3k7AAYFdoS/UPEGMA"
    "DACL/vltE7z5fqQtAqDWuiz/DVZbs4hpCivEn+HCzJ+g0AAYPUKXF84u0oxeTA6A4hFg/4ucjrziHdI2"
    "AUBfAh/JpbtG81ALBoM/vXkiPADQ29qkwLJaYggA0hGOPMs65jhtVQKgrwOyjUs8KQqKAHYCDAJEJpLw"
    "ABDoLNyQhV05zZQKIF0k3M3lzcJyWNsEgJzpAdUEyM0IiqUQNB3adC4zUYgAkBWD4IkdHEeOEEMADMzT"
    "iAD5ZDQqLi7zSNwWcLkAwKKLOKDKZQ+wz4rUPgS+xZPtwgRAoDtkNlyZKBcbAgA578MAuK5cuL3h9bOm"
    "A/H0S4gUILABgED/Ba15+TACdAD7DBK8BbIehQoAY//SmIER+YIhANBfNc2/ngx6fWtOcuUYTlLHRdkD"
    "4FgBYARg/Cp0h+x2KILo0urQABBI/eCYeNTV9ew/lAuYssm2h6+IaIZDxeu6Sl+vrypsuU5sLWqxtha1"
    "X7VlSaeY+kwnOQSsqADkTIF0AM6aClz0zblQHhmPqw0HANyReAB4B6E+qfUE1gNjhJMdECEOuAaGQzTd"
    "Vu8+n48/vS5KkhlybMUIJOjabDCa+qn/Rc4qeFTSPSQ0AFi8a5ehbpXaKvH/+Uc0k1UBrNo1OR8uPm8a"
    "X1P/SUykFQClA67LA0/8ehC46EdwHcFqFCoAygqIoUPayV1T2hmxARja7NqYD7TwewsEC+ZE2gFAiJH+"
    "RQGc4525vqMCSkaCZQR9KlioABDvnCn2mE7YyscdEwA4aG/E+IQoyl2JB03LRU6Q9jIrADDHeRFzoEXM"
    "/YXv+PpfwZUYQ/NDBiB+bOQNYzQWOs163gQASte8io1PCaUKAY57vnEfCpBRsgIAcZxH/R0c72E7+ulP"
    "tg9+oH8SEjIA8YzFy1oRsIn4WNtJj4mLsq7SxvKOHqS79gqScDCLFQCQcIp3ATdG+YwKnyu6Z4Au8c4c"
    "hkIGwLOjndgQzQZzOd+ABEDqez6idqYEegNIO5SgRATYAUAZIL/nfpU1GNOI/DdpHQZeHpFQO8EpFD4A"
    "HgOkvRaAypzWe5vKsOi+GbYGRWPuzrsOEX1cwA4Ayp1e6ul73+pOur6VbcLQVYATqHb6PyEBILkHQENM"
    "veHfDzTJGszXklbpulQAcDJtHRaAXYzHFBH9mClbAJCSSuz6KwFeU0yf6W4AAB0lSURBVNsqcFAyBLIB"
    "adaENADQsL4zGNxb9QmRTBQjlw4Acp6QBZaGxWIlhrswLQF2APByqh/ICCQ9t7a6iZ/CPsAmkSYAlISf"
    "rHkxXH4ZFQYkgLKYnsLcmiEgaZWWAGsAiB9DI98pH9+i1Z6ROhCA7JtIAG/WjgQQ6N4CrtnbyY4RAEjt"
    "YelSWwOjMa5imhSRJQAIExU1wAQczL/WvYUGiJaBw2F7DAjyfZYAIMspGKnDHA9+voVwSgkA9Y25IyW3"
    "NjIcy+9qqvGtAUDYhxKq8aTNG8WZcwMHe4oXo/QBoAh4UUYwGG+RH7KYAQmg1oBKbI8AmdcePjrSkhHo"
    "beNbbgRCsr45LGgbnAjqjFDaVIC6teEZcE5AKYHjyAwA0RUnlcFmSwfQ2WvgQjw7EsB7q91mA0FOTEc2"
    "1wHsMOgI9ViSTgBQFBXBWVnl5mTc1ywYcAmnh/c/yLmts8M4XpQPpYhtAeC9tE3gwCi3eY/dGrg2bj9K"
    "KwCOw6bDb4Vifi27OBiQ+uHRBBX/zIuTWiFAfepEqBbTGgBq/QYOcFzd7A8LweemN6YXAI/OUs1PS+Z1"
    "JUYkgBd5PN2Lc26HAFcuReEDkNsT2A6Ml1382ewgaAAVHPESWmkEgETRJFgE4Ii8VxiRAPFXsWOstGQJ"
    "ROQioGzYJgBkN3DiifIPl1+kOweWgqLiaBSlFwCv0LKMazKv8jQyBIC3I4Vz4k3I5q0ALiej0AFAD0O9"
    "9i4+R7765UzOBlukbmHpBgAJZ09Ek3bFty0wBQDxStE+6imleQKoi2sWWD81rNVnzyoFAOB041cSgP0J"
    "DgOuRGkHgDjo5/ArwS6+aPLWJQEQFwJMTJ1YDsceLsUKOBc+AOhzwA/ALv/q/Jqp5aAGWBaN5+XSCgBi"
    "Tk4NHAxQ6ux3xgDw/A6GOhWOzWtMORpzDNVt7sz1swKsqgDWGQSA3tEkAX4Anpzo/oVdBgAoWXVCe7rz"
    "xiY/xQAAnr8uxLptG2ZQapQA1y0MHQAyqzdQ6cnx2MadJugUHMLaQgKOzLUKgHLRZupmi+K6C3dpAIBG"
    "lgganjlp7cHz5b0iPstNIXvI+XmBSLgAIHQYAEDRfcZpUHpX1oB6K6MEXQYAeCmBNdfrmrjm7DAIQLMv"
    "Lq4d3qlbq5U7q2vL9cDbL2/onUgGVIcPQJUEi0M7N9YNvAtnsbypbJcDAEoJXCFhAgrkzMasoGkAklsT"
    "zpXrEfiChA7AqF4gAM95fiBh4v/snX+MVcUVx8/Mnblz774CwsouLMVlw4YV+bG7FEE2WZDfLL9/VVgL"
    "7aItv6SAWGGhiC3QSmkNFpTUgkUiu61ECRWIpqEUDTVISbW1NkVLYtTaRpqmhtqGajKdebtLgX1z373v"
    "3jf3vmS+/mMIvDt35nPPzDkzc84p9W7bafCbMzm/AIhmVP1HXcWVcrqybbstXgAI+9YsT09g8TjAmgGA"
    "ryuSxyPeIC+KEzyuq/LLmlHrO1ta3gHAhzzL+E7rheMHQOrNqer4Oyo+0dmg5hkA3F1ZzpbfknZ6FqgX"
    "ut+DBAEAN3usx93imxICAO42wYPT4UQ3AOTVIQoCLP6UBADuVRLrfOx7Bsg7AMJ0zq5XlzNDyN2UDACq"
    "oIs6twX/HdZuAfAR1VEfOqFCLK+HViu7tetkkhwABAEPWR5nQ9B78q/EDoBs/Uy1H/C10doBkNVkM9h/"
    "Wb8CnYRSWKfuVHlLI0kAwPPq+8viJeYnAYD0ZNWoXFbzsbq9AAwHOofRO2qYnIJyUC9b6Zv+x18DAEKP"
    "NXgcu+JzRidhCpBDusBRuis7dAMA2F6urGVbNwKeUydl3NUUoG6OFgDgA69YO7qYEACg/+tKAB7UDwD7"
    "GVcBwD+zu6nTZC1iiQOgZrWHl+3wAQkBALorFtaUH9YOANjr1ACcgl8q8wLxHUFS3+kBAM6qa31ySl8v"
    "rRmWCABOKMKWFG2JAYCaaUoAdo3+hXIJWNc7gQDYp7jjYQIu4amJAKC3IiRMpa+iGwAGf7zxvMdVANCX"
    "BioB2BgoYbomAMiP6zyWAdaE/V2DAyCTEbBrRCJIFH9OBcCRTq+XdwAIfl8JAC9TGlT3zxAk350uC8Ce"
    "8QgGID79zogsQDgG8LnMMxVFK5h+AMiIJTzoGTfEp65NIgDMrro9t8v8mQAY19z8hQ59e8qGq1u7r94R"
    "tl7IeBUAt+oHQFi4i4EBcPmWchwkVbYmAMRvfZTb3Y1MALS6dbe1q0fDNWsLdGRQSD+wK1csAhtjAADw"
    "gByqtX8xkAHQBgBhor+siAC4mbu0Q9cN2ITZ4fp85GAVANP1AyDUc1dgAAY3A0siAALonrssGg0AR2UO"
    "l6ulnf8vxN8P1+HfVMUB0FOxAMAWBe0x+rxnkbg4AQA4zp1oAJh5AwBX/4+3hOvwbUoALsQCAPwr8Cfz"
    "MASrl6IRAJw6xlEkALzIHZRJFtozNEx3j5nBVcnal8UDwN23Beyu4ikBnWGdFgAmDYwGgH5WZgCEDejL"
    "QnT3a1SVcAt1iQWAFDsY0Al8IGjcTCcAmGwtQ1YEAFQMpgoAUHXvHCuGYcZKliAlABviAADb+GSgznL4"
    "TGDJBQBg6ArqRAAA+4FFMwWO5B+tJyQXAuQt88PK/TU0rQbHAQDG9wSaAxw6NujbawUAwz6XRgAAHHXc"
    "jJFDYQP4MmwznMP4s5+oT62gRzvHGHVMATgFWwL11qrAOyd6LQCRJxkjAOAsp5lDx4g6qI+MPOOgxjb1"
    "mjrLiGX10X4quB0A3DfQFzMcSMIBGL04YEAw4xQgFseqFBrCF5yXkrcEfayGO6rGYRsm7/UgE/Ht8QAg"
    "XLp7GgL0FzoQeP7TCwDGRXeh8ADYcEweKFacm7d443a5qCM4W9U4kq5yJBcNv7/d8licoFsJQCwAiHf4"
    "foDO+pwd2AnSC4Asi/UPHgEAr3BlZjDpDDoHBzT53RXHMOLTbG26BDFZAGGb5gborA8CzwC6ARAN7FeJ"
    "wgJAWM856tRw6ZAgH/VIy+4+PjT37V+Pytaggd/NAJMWAMTLv1Tmv7NOB98L1Q4AZi0oSDBAcR7ggkcC"
    "gHRECAXYSENZNti+nGJxASBWsw/4fo+JOTxANwCYwIiJiIYGYAP1uHXMOzaJqB/JvQQvOijfB7FZAEJg"
    "oW+L+ZcCAECuvH/Iw1sAeIt7jxtH/uX5M8X8WZvpvx5+dQ5gL/lNfIQ2FwIAMubyFg8PQDeH5y1R/A1B"
    "gHcyric1ASB8lG/4bGl9bYEAgJ+o8x8OUgBgw3qEkA4C0HRg2nMEXXeS5t8+G5rLDBAPAGS3/9y+SgDW"
    "DkM6TADiCzJXEdQHwBmfTT1bGADIs272s2GnAMLwfD0m4LLiqLEmADAhpct9NXTPiAIBAJgN64qtUADI"
    "OwHla3SYgPrZEEOq2Oui1WSer1XzX1mhAAAkBRfKUCgLIP3J+3sgmlcCkMUrV6pDWlosgKwhs9lXTOMN"
    "XDAA4BQZ0zWkFyBnkhctauV1/Kn181ocQ7r4G/yA8tVZ20rptamXkw4AkFI46TNS51k1bDi38jgLCAOw"
    "plx501obAMLWtWR9SYc/AgUEADCWOu9v6DyvhpHLWcJ4IQkY30t91UgbAMIEfJz1Y7H43wsKAEzYE1PD"
    "WwBcujFvBIjJpe5dj1MF+gDAZPKMrO/YY3RBASDcuEFbrdAWoJ2AvIw/teZs8rplpQ8AYQJ+lNVYHQy+"
    "ExwrAMKNq92JwgKQJiAv4QCL8jv7gdctK40AYFiQtcEPQWlBASBMAOlCQwMge2dbfub/iXdgzHAiAAA2"
    "Zkm2IwsjczwOHxsAmNjkshseAPHQuziPto68/K3zI9IpmRMCALRmafELuT4yCgCQJQGAoAgSaH4ShbcA"
    "4rGPPR1pvSgBk/VgOaRwlubrAwDjT73fD82FwgNA9GHfKAAQH0jtNmpFWTXu3OPZT5brXAMQuHuOZ4PL"
    "7i9EAAjD0yMBoBxg/3vIjcIIUBehyr8NZdkvFui0AOJhxzy/lSuA4wSA5mgBhPEegrLMxn5zBNnHz/EI"
    "/AHxC63fSRHInm1IKwBMVpP1aPYzsQJgUTQvRwDwXu+jIdQvAOL5/V9pRDwcAshqaH0Z2m4X4kRZALx2"
    "lFfDX44XACdHAMS/mLza42wvQi7d6tMCEBts+8T6J2k6n7aPw54dT7k2w8jOhSNt6aCCr1tOGgEQjTrv"
    "8RaNOY9/zAAAYzsyBfLabLkEgA8Pdt1x8qFZK3KrHjlsTfczNR2lx3zdcdEKAPmkuN3cusWdtDAUAOKD"
    "cTJUWOuCO71GlQRAfmDOdf+g2M0ZAPGPWl3xTterbEh9dX19fXV19fKlDwcBIJ05supPb3S/98r4OQ2W"
    "60PiIdU7j61fdmIMuza24Md63eRWZv7NPACAce2GflMmCU1pGlNyo0pD/HJPhUr6d043l3qul+qvBw4E"
    "XVVNU1NJr+tV0rOoqKampiitIBkP5PGZ9LVQYg+qHVPRy4fEE/qncovPliilMbutUUYMcPissT5WsGqZ"
    "QTAyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIy"
    "MjICZhNibqHHIZKMZkxav/fSfKleLGUg0KC1t3w4V+qzmmS0p2JxezKiFjDjr0MvcDedRWxNUrr7OG+r"
    "KE27gUEg36qC36TzYrnI/WlCmmSzLe0mYLoBQIPBrZdJ65CDWnAqGS1Ksf1We1GDPgaAvGtWWz16Xn1f"
    "UpbdhMFRmUYXUWdpMzAzRHnVaUcmH6WIfwKlCelr4QBWzJA5dpFjPQo2JsYK5KmjxX9Fq2RpOkStK8lx"
    "A8UqAP8znWRZ2IB3xIxgAMjbbAu/5ekM1FbZhkShyWBNe3r7GRUYEzNS+TIB24sdS84AfF7S4m6TBrfn"
    "U/8DsY0JyNf4V/1XfP+WWG1PLEraZ4Z/1ZZR23JWAjEmIC9dTPBWR86zlLqHIGkWgNU0tida73ofNiYg"
    "P9pUbInxd11+GBJnADB8JIvkWNThrWAbE5APVa1KR1uotackgT1M2NtO2yTAT5YaTzB6+w+wl7ti/e9S"
    "fhZYAgEgJTvTNVEQWvouNhHB6B2tfRYvFmtAl88SfncS+xef6SittYaZTYHIe7fnallBXFjY5b2TaADS"
    "RuDzlKYnAdTdABD5DLCRuzLSRtFmSGoFEJyaLiGV/uBXzJhFGwFg87msI0Yt3gIMEnv8qt+EtvJa/OnZ"
    "xgREaVvx2EpuCdNqOTuLCElu3+Ldle0EHDPbglFq3HhXfv0WGnxAFpJM7sdlt6brZlucfgjGF4zuw7oo"
    "JgDRs3RgH7nZluSO7b2nfa3qdMFgvMFoFgCwTLhWchOILwKAhB+/PlFptdW8/2ozEDBnxaMA4HGrbV61"
    "pjX9j70rAa6qOsNnueee+3gQa15J6MMhQIIxgWDAgbA0gQhYCIhQqbKHtZS17JZNRBAojDJgWYUqS4fK"
    "1CoUGDYHqFbBpXSQDqBF244tjJ2OrS1j7VT/nv+c+x6LLFlY8u49/xAmw5J55/zfvy+npgOAOU5PowIA"
    "dtdSVsAmhatJypVaN9K80sojx2u+AlC0yDSHcLpGfVaLgGonAHJ7AD4/7PHI5tSYv2ldGyI6HSCaSMcO"
    "DFWT/4ydVrzHJgDxIUmN63QX5MSM0fJmoxGwCKhGAkCSFUII7AKBqUUsVeRptXYClNcytKsdGqymAdgS"
    "41y3AebckSIKAB3B+dppBQ9qZ9iEULX437Ut3iQVnnc+ZfjPHJn7QzBVAf5Olm0RrLoBYCVdwMMEkIBz"
    "6P6nyE06UdLrBW7yQfAfErU6oKqyVPhfITQA6OAsljoAQNW1IM+Mi3H4m8yyoWDVVCl5WVBPF9g71WUp"
    "lVpXCFiM40vqs8f4H9xMi4Aq0SoQnnam2zassT0AV7MCDnkXDAKALrYIqLwIqa/NgGOAWAOuk3KVNUwA"
    "nqDc9IgVfEKsFajs/WVFV0dA6IFL+khN7QG7JoJZ5m+4QQBf25BYHVA5DeqSIxFl/3V3xQAmU/MIT7TS"
    "2QDu8bR5LNMGg5Uh+WieLqopB3BvZkrm0pgjSUmaaRQXMGGdE7U6oBIOQGk7wA5bLyZeap6qouM4zta1"
    "GgFcwJIMOzJYifhv5iB1aVwBAD7vkMIHkbK8XUIHtK9LXGsFKuhAl3Y0ExYc0no5qVtNUx+cHcnhujKY"
    "A+MaWQRUjNxhQ3lca0466KnUdp+V+/qmnhMAdaL2jezUaIWkZlh3jP7RAnRen+odNTLqruA6H2AQYP2A"
    "65vNH72lW6oEpc3KAzBhxVgTjgmNhA6wVuDawbN8rFhXUtVX5AMmU15gHHWExoAVDfAiFgHXp8NtcQsg"
    "5UJ4Y0ggKunMYQP8NlGLgOtSw3ZcoAegIsDfuU4gCunYE2YQAF6c3nufnRi6hv93vCCG3Fe/UP4Dc1MJ"
    "HaD8ALh3oW0TvZq1JM+VxTxq+L+YBOmeLiCAQ+/X7QaZq8T/b3Ol/3XYjLvWAtVMm7QC6vdmsxw7M3gl"
    "+hIi2AAMaP9JwJqp1Wk+TSIAVtuZwa/fD9uOK4Ax/KMF7znB66Z39Iirv1b+SWJbhS+7ntyzegcwCkjZ"
    "LscJ4EiVw1ZzMKWhGD0dtVuELqF6U3hE6KQ5pB/CNeABnKdxyBhBjSvoiR1D7NwoIf4GBYc8NQHnKTUA"
    "hu4hMpivsDHJjuebLjHuwecrlRkIuSuA7McdCmxWep5esCYEjL6PSBnQgTonk5R290yjE4W2z5EoCTME"
    "/LMzV/YRuAFMASAS69HBpM5YUM9ctM+khdVx8+YSFnXDiwDNZSajLdZQtP74EBQMbIHxQIDFgpHvvAOm"
    "XZxHYFs2y5ShzQpp/kdJyW6hnwHiSv9vl8F0/y46tEOyf2yKg+rE0H6/MnhhjQiRz1Fn/TIc/8RngChd"
    "QYz8B/rYkrB39cAbrpan+bMJCS0CcNxjlFL8mP6nnOYt1wtBAm8T1fn+XmAel8DHL34bDS0ApGy+Fzw9"
    "/SciotNWP/wP/G0omJc+m+O3vcZgaX8Z1vnxBaOxS0ZHAHRHPeaEBgCZrP8XEOcmJQRps8OYEGBZztEc"
    "iEQ8BADAyah0w+MOMymbngR/t7yCwKrMsO0UVYCv+ydOY7gAFrt/z8twLdVUCCD/6Mb1/Dj3OLQa5oYM"
    "AXJLmoeRn14A2/0NJ2xWECcHD7+lFyDqPfidH5EB6ICtOD3xPgJfG39KF2UQGbq6iNIBbMhA4CYjoC5i"
    "wzoSEjWozli6LwejfxMKHYyGMxJ21KHrm4fGMBaG/MV6n1gYEDDNA5MKUbRsV4irIVG2YAIIT5jxUXi5"
    "lxKFYDcNO65DSpcAvv1maPAQ5oY1CsaFgqTWGeUJoTOIwxDfP+/IoNcGcqc1A9P5wQXk/JXo/siQFkUV"
    "AhS3F+dT3SeC7fCRjZOIDOhlYIpHyvJ9mPfTm5Q5nzqTmO0v4S2JKolnbZb67w1STiNlPSULqBVQXB67"
    "Dad+je9P6UPZIfT+rxANOO7cPBrj4F/M6EMOiQZyjlweGKk7IUzxP+1R2xlrMgLKE+g6hSpHQAOAAx1e"
    "xKQbLNlg6Px9AVzQBADOdNBxkCVUjS7J7VsG8YRqhI7fbR4w5chI3dMx6pnCTxzgwdckc+2AVMIPwG7B"
    "PeNVcJSQD9qlTmCqAxjsuO7RQeAJv/AH8Jd6tiX6UjOglGG0TzrEIgkdCRu3Big2+moJNU0fmmoftxPS"
    "lytIXDFPimZAPO4ZT4BDZHhRQJTAG4/jzjct/eq3+J0t7KKcKykBNAV3d9H5ALMiOdbu1UYBCP3anBUA"
    "Sc1Ge5Qy0/lheX45AJAy7opjn4AmIWDokx1YaqcGh00vA+oJX/lD8U4Me+wzSlcUFUPzNnAQyXiAj2hS"
    "mLqZUufU9jLwPGPVsOS1vVEI2n6r7gf439Q5plRmUmho8YohLDUrBCUP5wvwmz6QNnSVjNhX1K4HAOI2"
    "H7WWXpQxpSP61sXNwyy1jrPuoc6CR7yE8qf3LHccy/sK0tj3m4G4EBLS3ud6ydTSAguH5wH4rj9SWv1a"
    "xKZ+K2o6XcImfVshIJkXAig7sRDHi1ND+jPLF3XDUV/BqQFA+oAhJNsWfioVFLAFgyGuEwI+BLwPdzmX"
    "uIs1lt4cj5LPkwDI+3l/y9Gq0Podul2QJ5xoWNKgNVZQajQETh0sBhpTrr+HjS4KuJ1PfESiVvdXjSZv"
    "FJD0o7B3bOT8PQ6puRv0mm46U6axKoTpc6J5w+fZVSjVoX8+jrMTF3wBGnnm7dYqUojWQAxkNOiHjV6a"
    "9EcGb9tC1+7FrK4heBpQBwi/RiAARr5ySLpuVNakjJrM2nI2HbTtN7xXlH9uErFln+pGBFEVUp3pDMIz"
    "j6jrVdq04MUv/0ikrDHD1HLrtE5+U5sh7kHvxhOJa9M+N8izenUQgD9HoUtqFDo/M6d/lobI7ZIx1O1O"
    "VJJo0dwXI5T6G9DQb1V66tjRDliesqHfjaLW3+tntuiaS/ZiHtC2gw+U4I6122VlHVehYP/qpTrmT3wy"
    "bQKWvpeFTQ6W/TdO2hjJOv57T68X8x+mx++g5cYGbTLlbYPAR3N3l5n3kbWDii2/QJudfkBPudmq/w2n"
    "lY0f9BNsZtOc8gwA8sYf3JSlfC0ZvSWNpAzBhpa9xfqH+8WU5o9dCFK8GPB9ozJ83lvjf8PvnpDsMYML"
    "FN9jicFKPV0M0P2zxROlVsquvIkVF818zdj95z8rjvlvY/uqH8GYP/1Qlv+vLN0sEEza3IliUCASPiGA"
    "unpo+edVR/q7RM+aOGbZ8o1lA3PMXr/CsR+80q/AS/IdN3xj5QpyXpozFv+BY/l/c8llk+en4Xytx2lS"
    "+gRAnLYc//FrJbhtxDHEElRNucfZBR1zFrb59+lW3Tzw5zu4sUY420yP3bmQ6cWXVvXfkqDg7uEjKG4Y"
    "utgU6MfpWi45M2fm/SzBOuZUQx0k/zcqlomvr/jmPXG90t9P+PCEDYj/4NzzuZbxtzYqYEPGDFyG9j9C"
    "4RLCBtzePb4x55NvNU0snGCV5brUlNDlTdeV1//XuBH+lmNd5hPIfB7TL3wUn9yVa5l/G9IwhNSbta27"
    "foKEXgYCnDP00qfOmLZzQVGhD4AEY6/Wl3OFv5eFRZsOfPz0C+n0gsX3ma9JgKj9s5/UkpYbtwcCuE3i"
    "F8/f9ctuAl+l5JdCwDy+F4N2E6YM//TX5QsbZVfiRxeeOvzV6oN7ezzbDcwql2Qx6mJqubvnY1mWD7ef"
    "2syZsYx/jT1+2tj/85yOE8ZtmD6tz/+WT35gf0nrppdnDFh23V4rSycfOf/TVWuWti8u83jCxUz+qEt+"
    "vugyfecpYt8+qimUfUffX/2/vXsJhSCOAzi+ZnetMVxsTHmz5FEU5bHhJpRw8CaytdyUcLG1eRQhIgdy"
    "prSUMwcXByIHubi5ICevkjMzOx5rlnWzq+/nNtPU1P/3///+z2aceWruf50RmI1+asN7KMWMylWnM8pT"
    "pPEUzbZXdK3aC60R5jA/dcj3Uuly6uaTJ7MMhD9YmCTlxI0l66x+3iFqgTJ7P8Edpm/DfrO5Puz6Z9RL"
    "sa581FVgsAimFFZ6g49kq1q5zrZbvftFVlFJBkZdIw5MlwHUu31Dy087StwFSQ7niHcwDgrVeZslbfBo"
    "s8wuaqnA6DewgakLPcoCkzKW7PU0bC3GycpbwmX5n/7U6T8tEghS9cFh4sVwX/x3Lf/H7P+W81/neY65"
    "m7bbKZskeP/w8RklHfSDAzm2JnrpvNtTkfPrLkB7ymqv3R7Za3msKdbq1JfgUwFCIhEYlH1Zk8EyVpK6"
    "/jAxs3tXduyIifi2+cfnLkT1n+xe1e9fbgzkqx2K8AOKOARFFg+MT7kuW6czmxOa7tM7Gxuf19xJze7T"
    "1ujUydKOHtvHEX4hEEozpPOD6ZMvIRV+h0IM0Z7hbbPHh+/+McEFAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAADA33gBSI3UdwGELXkAAAAASUVORK5CYII="
)

# ── Core logic (no file I/O — returns CSV bytes) ──────────────────────────────

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def normalize_email(raw: str):
    email = re.sub(r'\s+', '', raw)
    return email if _EMAIL_RE.match(email) else None


def normalize_phone(raw: str):
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    return None


def get_fieldnames(file_bytes: bytes, filename: str) -> list:
    if filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        return [str(c) if c is not None else "" for c in first_row]
    text = file_bytes.decode("utf-8-sig")
    return csv.DictReader(io.StringIO(text)).fieldnames or []


def iter_rows(file_bytes: bytes, filename: str, fieldnames: list):
    if filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield {fieldnames[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(fieldnames)}
        wb.close()
    else:
        text = file_bytes.decode("utf-8-sig")
        yield from csv.DictReader(io.StringIO(text))


def process_batch(file_bytes, filename, phone_col, name_format,
                  name_col, first_name_col, last_name_col, extra_cols, dedup):
    seen_phones, seen_names = set(), set()
    rows, skipped, failed = [], 0, []
    out_fieldnames = ["name", "phoneNumber"] + extra_cols
    fieldnames = get_fieldnames(file_bytes, filename)

    for row in iter_rows(file_bytes, filename, fieldnames):
        if name_format == "first_last":
            name = f"{row.get(first_name_col,'').strip()} {row.get(last_name_col,'').strip()}".strip()
        else:
            name = row.get(name_col, "").strip()

        phone = normalize_phone(row.get(phone_col, "").strip())
        if phone is None:
            failed.append(name or "(no name)")
            skipped += 1
            continue
        if dedup == "phones" and phone in seen_phones:
            skipped += 1
            continue
        if dedup == "names" and name in seen_names:
            skipped += 1
            continue

        seen_phones.add(phone)
        seen_names.add(name)
        out = {"name": name, "phoneNumber": phone}
        for col in extra_cols:
            out[col] = row.get(col, "").strip()
        rows.append(out)

    if not rows:
        return 0, skipped, failed, None

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=out_fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return len(rows), skipped, failed, buf.getvalue().encode()


def process_email(file_bytes, filename, email_col, name_format,
                  first_name_col, last_name_col, full_name_col, dedup):
    seen_emails = set()
    rows, skipped, failed = [], 0, []

    if name_format == "split" and full_name_col:
        out_fieldnames = ["email", "firstName", "lastName"]
    else:
        out_fieldnames = ["email"]
        if first_name_col:
            out_fieldnames.append("firstName")
        if last_name_col:
            out_fieldnames.append("lastName")

    fieldnames = get_fieldnames(file_bytes, filename)
    for row in iter_rows(file_bytes, filename, fieldnames):
        email = normalize_email(row.get(email_col, ""))
        if email is None:
            failed.append(row.get(email_col, "").strip() or "(empty)")
            skipped += 1
            continue
        if dedup == "emails" and email in seen_emails:
            skipped += 1
            continue
        if name_format == "split" and full_name_col:
            full_name = row.get(full_name_col, "").strip()
            parts = full_name.split(None, 1)
            if len(parts) < 2:
                failed.append(full_name or "(no name)")
                skipped += 1
                continue
            first, last = parts[0], parts[1]

        seen_emails.add(email)
        out = {"email": email}
        if name_format == "split" and full_name_col:
            out["firstName"] = first
            out["lastName"] = last
        else:
            if first_name_col:
                out["firstName"] = row.get(first_name_col, "").strip()
            if last_name_col:
                out["lastName"] = row.get(last_name_col, "").strip()
        rows.append(out)

    if not rows:
        return 0, skipped, failed, None

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=out_fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return len(rows), skipped, failed, buf.getvalue().encode()


def process_contact(file_bytes, filename, email_col, name_format,
                    full_name_col, first_name_col, last_name_col, extra_cols, dedup):
    seen_emails = set()
    rows, skipped, failed = [], 0, []

    out_fieldnames = ["email"]
    if name_format == "split" and full_name_col:
        out_fieldnames += ["firstName", "lastName"]
    else:
        if first_name_col:
            out_fieldnames.append("firstName")
        if last_name_col:
            out_fieldnames.append("lastName")
    out_fieldnames += extra_cols

    fieldnames = get_fieldnames(file_bytes, filename)
    for row in iter_rows(file_bytes, filename, fieldnames):
        email = normalize_email(row.get(email_col, ""))
        if email is None:
            failed.append(row.get(email_col, "").strip() or "(empty)")
            skipped += 1
            continue
        if dedup == "emails" and email in seen_emails:
            skipped += 1
            continue
        if name_format == "split" and full_name_col:
            full_name = row.get(full_name_col, "").strip()
            parts = full_name.split(None, 1)
            if len(parts) < 2:
                failed.append(full_name or "(no name)")
                skipped += 1
                continue
            first, last = parts[0], parts[1]

        seen_emails.add(email)
        out = {"email": email}
        if name_format == "split" and full_name_col:
            out["firstName"] = first
            out["lastName"] = last
        else:
            if first_name_col:
                out["firstName"] = row.get(first_name_col, "").strip()
            if last_name_col:
                out["lastName"] = row.get(last_name_col, "").strip()
        for col in extra_cols:
            out[col] = row.get(col, "").strip()
        rows.append(out)

    if not rows:
        return 0, skipped, failed, None

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=out_fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return len(rows), skipped, failed, buf.getvalue().encode()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

st.set_page_config(page_title="VINSI CSV Formatter", page_icon="📋", layout="centered")

st.image(f"data:image/png;base64,{LOGO_B64}", width=180)
st.title("VINSI CSV Formatter")

mode = st.radio(
    "What would you like to create?",
    ["Batch Phone Import", "Email Campaign", "Batch Contact Import"],
    horizontal=True,
)

uploaded = st.file_uploader("Upload your LineLeader export", type=["xlsx", "csv"])

if uploaded:
    file_bytes = uploaded.read()
    filename = uploaded.name
    try:
        columns = get_fieldnames(file_bytes, filename)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        st.stop()

    if not columns:
        st.error("No columns found in the uploaded file.")
        st.stop()

    st.divider()

    # ── Batch Phone Import ────────────────────────────────────────────────────
    if mode == "Batch Phone Import":
        col1, col2 = st.columns(2)
        with col1:
            name_format = st.radio("Name format", ["Full Name Column", "Separate First / Last Name Columns"])
            phone_col = st.selectbox("Phone number column", columns)
        with col2:
            if name_format == "Full Name Column":
                name_col = st.selectbox("Name column", columns)
                first_name_col = last_name_col = ""
            else:
                name_col = ""
                first_name_col = st.selectbox("First name column", columns)
                last_name_col = st.selectbox("Last name column", columns)

        extra_cols = st.multiselect("Extra columns to include (optional)", columns)
        dedup = st.selectbox("Remove Duplicates By", ["None", "Phone numbers", "Names"])
        dedup_key = {"None": "none", "Phone numbers": "phones", "Names": "names"}[dedup]
        nf_key = "first_last" if name_format == "Separate First / Last Name Columns" else "full"

        if st.button("Convert", type="primary"):
            count, skipped, failed, csv_bytes = process_batch(
                file_bytes, filename, phone_col, nf_key,
                name_col, first_name_col, last_name_col, extra_cols, dedup_key,
            )
            if csv_bytes is None:
                st.error("No valid phone numbers found. Nothing was converted.")
            else:
                st.success(f"✓ {count} rows converted, {skipped} skipped.")
                if failed:
                    with st.expander(f"Rows with invalid phone numbers ({len(failed)})"):
                        st.write(failed)
                st.download_button("Download CSV", csv_bytes, "vinsi_batch_calls.csv", "text/csv")

    # ── Email Campaign ────────────────────────────────────────────────────────
    elif mode == "Email Campaign":
        col1, col2 = st.columns(2)
        with col1:
            name_format = st.radio("Name format", ["Separate First / Last Name Columns", "Full Name Column (split)", "No name"])
            email_col = st.selectbox("Email column", columns)
        with col2:
            if name_format == "Separate First / Last Name Columns":
                first_name_col = st.selectbox("First name column", columns)
                last_name_col = st.selectbox("Last name column", columns)
                full_name_col = ""
            elif name_format == "Full Name Column (split)":
                full_name_col = st.selectbox("Full name column (will be split)", columns)
                first_name_col = last_name_col = ""
            else:
                first_name_col = last_name_col = full_name_col = ""

        dedup = st.selectbox("Remove Duplicates By", ["None", "Email addresses"])
        dedup_key = "emails" if dedup == "Email addresses" else "none"
        nf_key = {"Full Name Column (split)": "split", "Separate First / Last Name Columns": "separate", "No name": "none"}[name_format]

        if st.button("Convert", type="primary"):
            count, skipped, failed, csv_bytes = process_email(
                file_bytes, filename, email_col, nf_key,
                first_name_col, last_name_col, full_name_col, dedup_key,
            )
            if csv_bytes is None:
                st.error("No valid email addresses found. Nothing was converted.")
            else:
                st.success(f"✓ {count} rows converted, {skipped} skipped.")
                if failed:
                    with st.expander(f"Invalid/skipped emails ({len(failed)})"):
                        st.write(failed)
                st.download_button("Download CSV", csv_bytes, "email_campaign.csv", "text/csv")

    # ── Batch Contact Import ──────────────────────────────────────────────────
    else:
        col1, col2 = st.columns(2)
        with col1:
            name_format = st.radio("Name format", ["Separate First / Last Name Columns", "Full Name Column (split)", "No name"])
            email_col = st.selectbox("Email column", columns)
        with col2:
            if name_format == "Separate First / Last Name Columns":
                first_name_col = st.selectbox("First name column", columns)
                last_name_col = st.selectbox("Last name column", columns)
                full_name_col = ""
            elif name_format == "Full Name Column (split)":
                full_name_col = st.selectbox("Full name column (will be split)", columns)
                first_name_col = last_name_col = ""
            else:
                first_name_col = last_name_col = full_name_col = ""

        extra_cols = st.multiselect("Extra columns to include (optional)", columns)
        dedup = st.selectbox("Remove Duplicates By", ["None", "Email addresses"])
        dedup_key = "emails" if dedup == "Email addresses" else "none"
        nf_key = {"Full Name Column (split)": "split", "Separate First / Last Name Columns": "separate", "No name": "none"}[name_format]

        if st.button("Convert", type="primary"):
            count, skipped, failed, csv_bytes = process_contact(
                file_bytes, filename, email_col, nf_key,
                full_name_col, first_name_col, last_name_col, extra_cols, dedup_key,
            )
            if csv_bytes is None:
                st.error("No valid contacts found. Nothing was converted.")
            else:
                st.success(f"✓ {count} rows converted, {skipped} skipped.")
                if failed:
                    with st.expander(f"Invalid/skipped entries ({len(failed)})"):
                        st.write(failed)
                st.download_button("Download CSV", csv_bytes, "vinsi_contact_import.csv", "text/csv")
